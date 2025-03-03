import os
import pandas as pd
from openpyxl.utils import get_column_letter


def update_error_overview(mae, data_name, target_variable, model_config, error_overview_path):
    # Create the error overview directory if it does not exist.
    os.makedirs(os.path.dirname(error_overview_path), exist_ok=True)

    version = data_name.split('_')[-1]

    if not os.path.exists(error_overview_path):
        index = pd.MultiIndex.from_tuples([], names=["target_variable", "model_config"])
        df = pd.DataFrame(columns=[version], index=index)
    else:
        df = pd.read_excel(error_overview_path, index_col=[0, 1], engine='openpyxl')

    if version not in df.columns:
        df[version] = pd.NA

    def extract_version_num(col):
        if col.startswith('v'):
            try:
                return int(col[1:])
            except ValueError:
                return 0
        return 0

    sorted_columns = sorted(df.columns, key=extract_version_num, reverse=True)
    df = df[sorted_columns]

    # Update the row for the given target_variable and model_config.
    row_idx = (target_variable, model_config)
    df.loc[row_idx, version] = mae

    # Compute the best (minimum) error for non-'best' configurations.
    target_mask = df.index.get_level_values("target_variable") == target_variable
    non_best_mask = df.index.get_level_values("model_config") != 'best'
    subset = df[target_mask & non_best_mask][version]

    best_error = subset.min() if not subset.dropna().empty else pd.NA

    best_row = (target_variable, 'best')
    df.loc[best_row, version] = best_error

    df = df.sort_index()

    with pd.ExcelWriter(error_overview_path, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, merge_cells=True)

        worksheet = writer.sheets['Sheet1']
        worksheet.freeze_panes = "C1"

        for column_cells in worksheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            worksheet.column_dimensions[column].width = max_length + 2  # extra padding for clarity
